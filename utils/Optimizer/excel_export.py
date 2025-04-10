import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_machine_agenda_excel(schedule, output_excel_file="machine_agenda.xlsx"):
    """
    Creates an Excel file that represents the daily planner with machine schedules in a structured format.
    """
    # Initialize a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Planner"

    # Define time slots (1-minute intervals across 24 hours)
    time_slots = [datetime.strptime(f"{h:02}:{m:02}", "%H:%M") for h in range(24) for m in range(60)]
    time_strs = [t.strftime("%H:%M") for t in time_slots]

    # Get unique machine names in sorted order
    machines_order = sorted(set(entry["machine"] for entry in schedule))

    # Write column headers
    ws["A1"] = "Time"
    for col_idx, machine in enumerate(machines_order, start=2):
        ws.cell(row=1, column=col_idx, value=machine)

    # Fill time slots in first column
    for row_idx, time_str in enumerate(time_strs, start=2):
        ws.cell(row=row_idx, column=1, value=time_str)

    # Populate the machine schedule
    for entry in schedule:
        machine = entry["machine"]
        start_time = datetime.strptime(entry["start_time"], "%Y-%m-%d %H:%M")
        end_time = datetime.strptime(entry["end_time"], "%Y-%m-%d %H:%M")

        start_idx = time_strs.index(start_time.strftime("%H:%M")) + 2
        end_idx = time_strs.index(end_time.strftime("%H:%M")) + 2
        col_idx = machines_order.index(machine) + 2

        cell_value = f"{entry['patient_id']} ({entry['scan_type']})"

        for row in range(start_idx, end_idx):
            ws.cell(row=row, column=col_idx, value=cell_value)

    # Adjust column widths
    for col in range(1, len(machines_order) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Save Excel file
    wb.save(output_excel_file)
    print(f"Machine agenda saved as '{output_excel_file}'")
